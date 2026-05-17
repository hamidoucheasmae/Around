import re, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Pack definitions ──────────────────────────────────────────────────────────
PACKS = {
    'FREE':      {'label':'🆓 Free (2 Editions)',      'color':'D4EDDA','editions':['talking','friends']},
    'LOVE':      {'label':'💕 Love Journey · €2.99',   'color':'FAD7CC','editions':['talking','couples','engaged','married','stillus']},
    'DEEP':      {'label':'🌙 Deep & Real · €1.99',    'color':'D8CCE8','editions':['healing','overthinkers','betweenus','solo','masks']},
    'SOCIAL':    {'label':'👥 Social Circle · €2.49',  'color':'FBF0C0','editions':['friends','girlsnight','secretfiles','family','whoknows']},
    'ADULTS':    {'label':'🔞 Adults Only · €1.99',    'color':'F5C0C0','editions':['spicy']},
    'ALLACCESS': {'label':'👑 All Access · €5.99',     'color':'FDE8C8','editions':'__all__'},
}
PACK_COLORS = {k: v['color'] for k, v in PACKS.items()}

EDITION_PACK = {}
for pk, pd in PACKS.items():
    if pd['editions'] == '__all__': continue
    for e in pd['editions']:
        if e not in EDITION_PACK:
            EDITION_PACK[e] = pk

EDITION_NAMES = {
    'talking':'Talking Stage','couples':'Couples','married':'Married','friends':'Friends',
    'spicy':'18+ Unfiltered','stillus':'Still Us','engaged':'Engaged','secretfiles':'Secret Files',
    'family':'Family','betweenus':'Between Us','overthinkers':'Overthinkers','healing':'Healing',
    'solo':'Solo Edition','masks':'Masks','loveyourself':'Love Yourself','future':'Future',
    'redflags':'Red & Green Flags','girlsnight':'Girls Night','whoknows':'Who Knows Me Best?',
    'firstimpact':'First Impression','artist':'Artist Edition',
}
EDITION_AR = {
    'talking':'مرحلة التعرف','couples':'العلاقة','married':'المتزوجون','friends':'الأصدقاء',
    'spicy':'بدون تصفية','stillus':'مازلنا نحن','engaged':'المخطوبون','secretfiles':'الملفات السرية',
    'family':'العائلة','betweenus':'بيناتنا','overthinkers':'المكترين','healing':'الشفاء',
    'solo':'بوحدك','masks':'الأقنعة','loveyourself':'حب نفسك','future':'المستقبل',
    'redflags':'العلامات','girlsnight':'ليلة البنات','whoknows':'شكون كيعرفني؟',
    'firstimpact':'الانطباع الأول','artist':'نسخة الفنانين',
}
EDITION_CATS = {
    'friends':     ['soft_friends','real_friends','psych_friends','between_friends','power_friends','friendship_deep','challenges_friends'],
    'spicy':       ['spicy_part1','spicy_part2','spicy_part3','spicy_part4'],
    'talking':     ['soft_talking','real_talking','psych_talking','between_talking','power_talking','desires_talking','future_talking'],
    'couples':     ['soft_couples','real_couples','psych_couples','between_couples','power_couples','desires_couples','future_couples'],
    'stillus':     ['distance_su','stayed_su','unsaid_su','patterns_su','choosing_su','final_su'],
    'engaged':     ['soft_engaged','real_engaged','psych_engaged','between_engaged','power_engaged','promise_engaged','home_engaged'],
    'secretfiles': ['secrets_sf','laughs_sf','likely_sf','cringe_sf','chaos_sf'],
    'family':      ['energy_fam','chaos_fam','feels_fam','night_fam','forever_fam'],
    'betweenus':   ['between_bt','debates_bt','midnight_bt','real_bt'],
    'overthinkers':['mind_3am','insecurities_ot','mindvsheart_ot','healing_ot'],
    'healing':     ['whathurt_h','healingprocess_h','neversay_h','newversion_h','softness_h'],
    'solo':        ['whoami_s','thoughts_s','healing_s','future_s','selflove_s'],
    'masks':       ['showpeople_m','hiddenself_m','maskexists_m','exhaustion_m','unmasking_m'],
    'loveyourself':['seeing_ly','softenergy_ly','healself_ly','innerworld_ly','becoming_ly'],
    'future':      ['dreamlife_f','lovenext_f','fears_f','success_f','letters_f'],
    'redflags':    ['turnoffs_rf','greenflags_rf','toxic_rf','standards_rf','opinions_rf'],
    'girlsnight':  ['girlhood_gn','feminine_gn','love_gn','neversay_gn','forever_gn'],
    'whoknows':    ['littlethings_wk','mindread_wk','loveread_wk','realme_wk','memories_wk'],
    'firstimpact': ['howseeme_fi','perception_fi','relimpress_fi','truthvibe_fi'],
    'artist':      ['identity_ar','block_ar','emotions_ar','dreams_ar'],
    'married':     ['soft','real','psych','between','power','desires_married','future_married'],
}

# ── Smarter parser: extract category blocks by finding each key ───────────────
with open('/home/user/Around/categories.js', 'r', encoding='utf-8') as f:
    raw = f.read()

# Find all category keys and their start positions
key_positions = []
for m in re.finditer(r'^  (\w+)\s*:\s*\{', raw, re.MULTILINE):
    key_positions.append((m.group(1), m.start(), m.end()))

# For each key, extract its block by counting braces from the opening {
def extract_block(text, start_of_brace):
    depth = 0
    i = start_of_brace
    while i < len(text):
        if text[i] == '{':
            depth += 1
        elif text[i] == '}':
            depth -= 1
            if depth == 0:
                return text[start_of_brace:i+1]
        i += 1
    return text[start_of_brace:]

CATEGORIES = {}
for idx, (key, mstart, mend) in enumerate(key_positions):
    # find the opening brace of this category
    brace_start = mend - 1  # the { at end of the match
    block = extract_block(raw, brace_start)

    # Extract en and ar names (category-level, before qs:[)
    en_m = re.search(r"\ben\s*:\s*'([^']+)'", block)
    ar_m = re.search(r"\bar\s*:\s*'([^']+)'", block)
    cat_en = en_m.group(1) if en_m else key
    cat_ar = ar_m.group(1) if ar_m else ''

    # Extract the qs array content
    qs_m = re.search(r'\bqs\s*:\s*\[', block)
    if not qs_m:
        continue
    qs_start = qs_m.end() - 1  # position of [
    # find matching ]
    depth2 = 0
    i2 = qs_start
    while i2 < len(block):
        if block[i2] == '[': depth2 += 1
        elif block[i2] == ']':
            depth2 -= 1
            if depth2 == 0:
                qs_block = block[qs_start:i2+1]
                break
        i2 += 1
    else:
        qs_block = ''

    # Parse individual questions from qs_block
    questions = []
    # Find each {en:"...",ar:"..."...} question object
    q_pos = 0
    while True:
        q_start = qs_block.find('{en:', q_pos)
        if q_start == -1:
            break
        q_block = extract_block(qs_block, q_start)
        q_en_m = re.search(r'en\s*:\s*"([^"]+)"', q_block)
        q_ar_m = re.search(r'ar\s*:\s*"([^"]+)"', q_block)
        chal_m  = re.search(r'chal\s*:\s*\{en\s*:\s*"([^"]+)"\s*,\s*ar\s*:\s*"([^"]+)"\}', q_block)
        if q_en_m and q_ar_m:
            questions.append({
                'en':      q_en_m.group(1),
                'ar':      q_ar_m.group(1),
                'chal_en': chal_m.group(1) if chal_m else '',
                'chal_ar': chal_m.group(2) if chal_m else '',
            })
        q_pos = q_start + len(q_block)

    CATEGORIES[key] = {'en': cat_en, 'ar': cat_ar, 'qs': questions}

total_q = sum(len(v['qs']) for v in CATEGORIES.values())
print(f"Parsed {len(CATEGORIES)} categories, {total_q} questions")

# ── Build rows ────────────────────────────────────────────────────────────────
rows = []
missing = []
for edition_key, edition_name in EDITION_NAMES.items():
    pack_key   = EDITION_PACK.get(edition_key, 'ALLACCESS')
    pack_label = PACKS[pack_key]['label']
    cats = EDITION_CATS.get(edition_key, [])
    for cat_key in cats:
        cat = CATEGORIES.get(cat_key)
        if not cat:
            missing.append(cat_key)
            continue
        for i, q in enumerate(cat['qs'], 1):
            rows.append({
                'pack':        pack_label,
                'pack_key':    pack_key,
                'edition':     edition_name,
                'edition_ar':  EDITION_AR.get(edition_key,''),
                'category_en': cat['en'],
                'category_ar': cat['ar'],
                'cat_key':     cat_key,
                'q_num':       i,
                'q_en':        q['en'],
                'q_ar':        q['ar'],
                'chal_en':     q['chal_en'],
                'chal_ar':     q['chal_ar'],
            })

if missing:
    print(f"Missing categories: {set(missing)}")
print(f"Total question rows: {len(rows)}")

# ── Workbook ──────────────────────────────────────────────────────────────────
wb  = openpyxl.Workbook()
thin = Side(style='thin', color='CCCCCC')
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
hfill = PatternFill('solid', fgColor='1C0E0A')
hfont = Font(bold=True, color='F2DDD5', size=10)
ctr   = Alignment(horizontal='center', vertical='center', wrap_text=True)
top   = Alignment(vertical='top', wrap_text=True)
topc  = Alignment(horizontal='center', vertical='top')

def make_header(ws, cols):
    for ci, (h, w) in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment = hfill, hfont, ctr
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'

# ── Sheet 1: All Questions ────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'All Questions'
make_header(ws1, [
    ('Pack',28),('Edition (EN)',20),('Edition (AR)',18),
    ('Category (EN)',22),('Category (AR)',22),
    ('#',5),('Question (English)',55),('Question (Arabic)',55),
    ('Challenge (EN)',42),('Challenge (AR)',42)
])

for ri, row in enumerate(rows, 2):
    fill = PatternFill('solid', fgColor=PACK_COLORS[row['pack_key']])
    vals = [
        row['pack'], row['edition'], row['edition_ar'],
        row['category_en'], row['category_ar'],
        row['q_num'], row['q_en'], row['q_ar'],
        row['chal_en'], row['chal_ar']
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws1.cell(row=ri, column=ci, value=v)
        cell.fill   = fill
        cell.border = bdr
        cell.alignment = topc if ci in (7,8,9,10) else (topc if ci==6 else top)
        if ci == 6: cell.alignment = Alignment(horizontal='center', vertical='top')

ws1.auto_filter.ref = f"A1:J1"

# ── Sheet 2: Pack Overview ────────────────────────────────────────────────────
ws2 = wb.create_sheet('Packs Overview')
make_header(ws2,[('Pack',30),('Price',10),('# Editions',12),('# Questions',12),('Editions Included',60)])

pack_prices = {'FREE':'Free','LOVE':'€2.99','DEEP':'€1.99','SOCIAL':'€2.49','ADULTS':'€1.99','ALLACCESS':'€5.99'}
for ri, (pk, pd) in enumerate(PACKS.items(), 2):
    fill = PatternFill('solid', fgColor=PACK_COLORS[pk])
    eds  = list(EDITION_NAMES.keys()) if pd['editions']=='__all__' else pd['editions']
    q_ct = sum(
        len(CATEGORIES[c]['qs'])
        for e in eds for c in EDITION_CATS.get(e,[])
        if c in CATEGORIES
    )
    ed_list = ', '.join(EDITION_NAMES.get(e,e) for e in eds)
    for ci, v in enumerate([pd['label'], pack_prices[pk], len(eds), q_ct, ed_list], 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        cell.fill = fill; cell.border = bdr
        cell.alignment = Alignment(horizontal='center' if ci<5 else 'left', vertical='center', wrap_text=True)
    ws2.row_dimensions[ri].height = 30

# ── Sheet 3: Edition Summary ──────────────────────────────────────────────────
ws3 = wb.create_sheet('Editions')
make_header(ws3,[('#',6),('Edition',22),('Arabic',18),('Pack',28),('Price',10),('Categories',12),('Questions',12)])

for ri, (ek, en) in enumerate(EDITION_NAMES.items(), 2):
    pk   = EDITION_PACK.get(ek,'ALLACCESS')
    fill = PatternFill('solid', fgColor=PACK_COLORS[pk])
    cats = EDITION_CATS.get(ek,[])
    q_ct = sum(len(CATEGORIES[c]['qs']) for c in cats if c in CATEGORIES)
    price = {'FREE':'Free','LOVE':'€2.99','DEEP':'€1.99','SOCIAL':'€2.49','ADULTS':'€1.99','ALLACCESS':'€5.99'}[pk]
    for ci, v in enumerate([ri-1, en, EDITION_AR.get(ek,''), PACKS[pk]['label'], price, len(cats), q_ct], 1):
        cell = ws3.cell(row=ri, column=ci, value=v)
        cell.fill = fill; cell.border = bdr
        cell.alignment = Alignment(horizontal='center' if ci in (1,5,6,7) else 'left', vertical='center')
    ws3.row_dimensions[ri].height = 22

# ── Sheet 4: Per-Edition tabs ─────────────────────────────────────────────────
for edition_key, edition_name in EDITION_NAMES.items():
    pk    = EDITION_PACK.get(edition_key, 'ALLACCESS')
    color = PACK_COLORS[pk]
    safe  = re.sub(r'[\\/*?:\[\]]', '', edition_name)[:31]
    ws    = wb.create_sheet(safe)
    make_header(ws,[('Category (EN)',22),('Category (AR)',20),('#',5),('Question (English)',55),('Question (Arabic)',55),('Challenge (EN)',38),('Challenge (AR)',38)])

    ed_rows = [r for r in rows if r['edition']==edition_name]
    for ri, row in enumerate(ed_rows, 2):
        fill = PatternFill('solid', fgColor=color)
        vals = [row['category_en'], row['category_ar'], row['q_num'], row['q_en'], row['q_ar'], row['chal_en'], row['chal_ar']]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = fill; cell.border = bdr
            cell.alignment = Alignment(horizontal='center' if ci==3 else ('left' if ci>3 else 'left'), vertical='top', wrap_text=True)

# ── Save ──────────────────────────────────────────────────────────────────────
out = '/home/user/Around/AroundTheSeneya_Questions.xlsx'
wb.save(out)
print(f"Saved → {out}")
